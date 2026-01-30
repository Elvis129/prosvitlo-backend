#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('/tmp/khm_побут.xlsx')
ws = wb.active

print("Шукаємо вул. Лісогринівецька в діапазоні 1000-4000 рядків...")

for i, row in enumerate(ws.iter_rows(min_row=1000, max_row=4000, values_only=True), 1000):
    if row[1] and 'Лісогрин' in str(row[1]):
        print(f"\n✓ Рядок {i}: {row[0]} | {row[1]} | {row[2]}")
        
        # Виводимо наступні 5 рядків
        for j in range(1, 6):
            next_row = list(ws.iter_rows(min_row=i+j, max_row=i+j, values_only=True))[0]
            if next_row[0]:
                print(f"  +{j}: {next_row[0]} | {next_row[1] if len(next_row) > 1 else ''} | {next_row[2] if len(next_row) > 2 else ''}")
        break
