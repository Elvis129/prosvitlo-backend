"""
Модуль для парсингу Excel файлів з hoe.com.ua
Завантажує та обробляє файли з чергами відключень
"""

import logging
import requests
from typing import List, Dict, Optional
from io import BytesIO
import openpyxl
import re

logger = logging.getLogger(__name__)

BASE_URL = "https://hoe.com.ua"
MAIN_PAGE_URL = f"{BASE_URL}/page/pogodinni-vidkljuchennja"


def fetch_excel_links() -> List[str]:
    """
    Отримує посилання на всі Excel файли зі сторінки
    
    Returns:
        Список URL до Excel файлів
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }
        response = requests.get(MAIN_PAGE_URL, headers=headers, timeout=30)
        response.raise_for_status()
        
        # Шукаємо всі посилання на .xlsx файли
        excel_pattern = r'href="([^"]*\.xlsx?)"'
        matches = re.findall(excel_pattern, response.text)
        
        # Формуємо повні URL
        excel_links = []
        for match in matches:
            if match.startswith('http'):
                excel_links.append(match)
            else:
                excel_links.append(f"{BASE_URL}{match}")
        
        logger.info(f"Знайдено {len(excel_links)} Excel файлів")
        return excel_links
        
    except Exception as e:
        logger.error(f"Помилка при отриманні посилань: {e}")
        return []


def download_excel_file(url: str) -> Optional[BytesIO]:
    """
    Завантажує Excel файл
    
    Args:
        url: URL до Excel файлу
    
    Returns:
        BytesIO об'єкт з даними файлу або None
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        
        logger.info(f"Завантажено файл: {url}")
        return BytesIO(response.content)
        
    except Exception as e:
        logger.error(f"Помилка при завантаженні {url}: {e}")
        return None


def parse_excel_file(file_data: BytesIO, source_url: str) -> List[Dict[str, str]]:
    """
    Парсить Excel файл з даними про черги
    
    Args:
        file_data: BytesIO об'єкт з даними файлу
        source_url: URL джерела для логування
    
    Returns:
        Список словників з адресами та чергами
    """
    addresses = []
    
    try:
        workbook = openpyxl.load_workbook(file_data, data_only=True)
        
        # Обробляємо всі аркуші
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logger.info(f"Обробка аркушу: {sheet_name}")
            
            # Пропускаємо перший рядок (заголовки)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue
                
                try:
                    # Структура може бути різною, але зазвичай:
                    # Колонка A: Місто/Населений пункт
                    # Колонка B: Вулиця
                    # Колонка C: Будинки
                    # Колонка D: Черга
                    
                    city = str(row[0]).strip() if row[0] else ""
                    street = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    house_numbers = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    queue = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    
                    # Пропускаємо порожні рядки
                    if not city or not street or not house_numbers:
                        continue
                    
                    # Обробляємо діапазони будинків (наприклад: "1-10", "1,3,5", "парні", "непарні")
                    house_list = parse_house_numbers(house_numbers)
                    
                    for house in house_list:
                        address_data = {
                            "city": normalize_text(city),
                            "street": normalize_text(street),
                            "house_number": house,
                            "queue": normalize_text(queue) if queue else None,
                            "source_url": source_url
                        }
                        addresses.append(address_data)
                        
                except Exception as e:
                    logger.warning(f"Помилка при обробці рядка {row_idx}: {e}")
                    continue
        
        workbook.close()
        logger.info(f"З файлу отримано {len(addresses)} записів")
        
    except Exception as e:
        logger.error(f"Помилка при парсингу Excel файлу: {e}")
    
    return addresses


def parse_house_numbers(house_str: str) -> List[str]:
    """
    Парсить рядок з номерами будинків
    Обробляє різні формати: "1-10", "1,3,5", "1, 3, 5", "парні", "непарні"
    
    Args:
        house_str: Рядок з номерами будинків
    
    Returns:
        Список номерів будинків
    """
    house_str = str(house_str).strip()
    
    # Якщо це просто один номер
    if ',' not in house_str and '-' not in house_str:
        return [house_str]
    
    houses = []
    
    # Розбиваємо по комах
    parts = [p.strip() for p in house_str.split(',')]
    
    for part in parts:
        # Обробка діапазонів (наприклад: "1-10")
        if '-' in part and not part.replace('-', '').replace(' ', '').isalpha():
            try:
                range_parts = part.split('-')
                if len(range_parts) == 2:
                    start = int(re.search(r'\d+', range_parts[0]).group())
                    end = int(re.search(r'\d+', range_parts[1]).group())
                    
                    # Обмежуємо діапазон (максимум 100 будинків)
                    if end - start <= 100:
                        houses.extend([str(i) for i in range(start, end + 1)])
                    else:
                        # Якщо діапазон надто великий, залишаємо як є
                        houses.append(part)
            except:
                houses.append(part)
        else:
            houses.append(part)
    
    return houses if houses else [house_str]


def normalize_text(text: str) -> str:
    """
    Нормалізує текст: видаляє зайві пробіли, приводить до стандартного вигляду
    
    Args:
        text: Текст для нормалізації
    
    Returns:
        Нормалізований текст
    """
    if not text:
        return ""
    
    # Видаляємо зайві пробіли
    text = ' '.join(str(text).split())
    
    # Видаляємо "None"
    if text.lower() == 'none':
        return ""
    
    return text


def scrape_all_addresses() -> List[Dict[str, str]]:
    """
    Головна функція для завантаження та парсингу всіх Excel файлів
    
    Returns:
        Список всіх адрес з усіх файлів
    """
    logger.info("Початок парсингу Excel файлів з hoe.com.ua...")
    
    # Отримуємо посилання на файли
    excel_links = fetch_excel_links()
    
    if not excel_links:
        logger.warning("Не знайдено посилань на Excel файли")
        return []
    
    all_addresses = []
    
    # Завантажуємо та парсимо кожен файл
    for url in excel_links:
        logger.info(f"Обробка файлу: {url}")
        
        file_data = download_excel_file(url)
        if not file_data:
            continue
        
        addresses = parse_excel_file(file_data, url)
        all_addresses.extend(addresses)
    
    # Видаляємо дублікати
    unique_addresses = remove_duplicates(all_addresses)
    
    logger.info(f"Парсинг завершено. Отримано {len(unique_addresses)} унікальних адрес")
    return unique_addresses


def remove_duplicates(addresses: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Видаляє дублікати адрес
    
    Args:
        addresses: Список адрес
    
    Returns:
        Список унікальних адрес
    """
    unique = {}
    
    for addr in addresses:
        key = (
            addr.get('city', '').lower(),
            addr.get('street', '').lower(),
            addr.get('house_number', '').lower()
        )
        
        if key not in unique:
            unique[key] = addr
    
    result = list(unique.values())
    logger.info(f"Після видалення дублікатів: {len(result)} адрес")
    return result


# Для тестування
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    print("=== Тестування Excel парсера ===\n")
    
    addresses = scrape_all_addresses()
    
    print(f"\nЗнайдено {len(addresses)} адрес")
    
    if addresses:
        print("\nПерші 10 адрес:")
        for i, addr in enumerate(addresses[:10], 1):
            print(f"{i}. {addr['city']}, {addr['street']}, {addr['house_number']} - Черга {addr.get('queue', 'N/A')}")
